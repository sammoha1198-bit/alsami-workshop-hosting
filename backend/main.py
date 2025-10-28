# -*- coding: utf-8 -*-
"""
Alsami Workshop API (Final, Stable)
-----------------------------------
- SQLite تخزين + فهارس + WAL
- POST /api/sync/batch   : مزامنة دفعية (UPsert مرن)
- GET  /api/search/{key} : بحث موحّد عبر الجداول
- GET  /api/last3/engines /api/last3/generators
- POST /api/export/xlsx  : تصدير Excel ملوّن RTL (Arabic, RTL, colored)
- GET  /api/ping         : صحّة + أعداد الجداول
- GET  /api/debug/counts : أعداد مختصرة
"""
from __future__ import annotations

import io
import re
import json
import sqlite3
import datetime
from typing import List, Dict, Any, Optional

from fastapi import FastAPI, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

# ========= Excel (يظهر خطأ واضح إذا لم تُثبت openpyxl) =========
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore

APP_TITLE = "Alsami Workshop API"
APP_VER = "2.1.0"
DB_PATH = "alsami.db"

app = FastAPI(title=APP_TITLE, version=APP_VER)

# ========= CORS للواجهة (يدعم file:// و localhost) =========
ALLOWED_ORIGINS = [
    "*",                     # يسمح لأي أصل (مع allow_credentials=False)
    "https://alsami-app-cuop.onrender.com",                  # صفحات تفتح كـ file:// (Origin = null)
    "http://localhost:8000",
    "http://127.0.0.1:8000",
    "http://localhost",
    "http://127.0.0.1",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=False,               # يجب False إذا كان allow_origins يحتوي "*"
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],  # كي يقرأ المتصفح اسم الملف
)

# ===================== DB Helpers =====================
def db() -> sqlite3.Connection:
    con = sqlite3.connect(
        DB_PATH,
        check_same_thread=False,  # يسمح بالوصول متعدد الخيوط داخل FastAPI
        isolation_level=None,     # autocommit مع PRAGMA WAL
    )
    con.row_factory = sqlite3.Row
    # إعدادات تحسينية
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA synchronous=NORMAL;")
    con.execute("PRAGMA foreign_keys=ON;")
    return con

SCHEMAS: Dict[str, str] = {
    # Engines
    "eng_supply": (
        'id TEXT PRIMARY KEY, ts INTEGER, itemName TEXT, engineType TEXT, model TEXT, '
        'serial TEXT, prevSite TEXT, supDate TEXT, supplier TEXT, notes TEXT'
    ),
    "eng_issue": (
        'id TEXT PRIMARY KEY, ts INTEGER, serial TEXT, currSite TEXT, receiver TEXT, '
        'requester TEXT, issueDate TEXT, notes TEXT'
    ),
    "eng_rehab": (
        'id TEXT PRIMARY KEY, ts INTEGER, serial TEXT, rehabber TEXT, rehabType TEXT, '
        'rehabDate TEXT, notes TEXT'
    ),
    "eng_check": (
        'id TEXT PRIMARY KEY, ts INTEGER, serial TEXT, inspector TEXT, "desc" TEXT, '
        'checkDate TEXT, notes TEXT'
    ),
    "eng_upload": (
        'id TEXT PRIMARY KEY, ts INTEGER, serial TEXT, rehabUp TEXT, checkUp TEXT, '
        'rehabUpDate TEXT, checkUpDate TEXT, notes TEXT'
    ),
    "eng_lathe": (
        'id TEXT PRIMARY KEY, ts INTEGER, serial TEXT, lathe TEXT, latheDate TEXT, notes TEXT'
    ),
    "eng_pump": (
        'id TEXT PRIMARY KEY, ts INTEGER, serial TEXT, pumpSerial TEXT, pumpRehab TEXT, notes TEXT'
    ),
    "eng_electrical": (
        'id TEXT PRIMARY KEY, ts INTEGER, serial TEXT, etype TEXT, starter TEXT, '
        'alternator TEXT, edate TEXT'
    ),

    # Generators
    "gen_supply": (
        'id TEXT PRIMARY KEY, ts INTEGER, itemName TEXT, gType TEXT, model TEXT, code TEXT, '
        'prevSite TEXT, supDate TEXT, supplier TEXT, vendor TEXT, notes TEXT'
    ),
    "gen_issue": (
        'id TEXT PRIMARY KEY, ts INTEGER, itemName TEXT, code TEXT, issueDate TEXT, '
        'receiver TEXT, requester TEXT, currSite TEXT, notes TEXT'
    ),
    "gen_inspect": (
        'id TEXT PRIMARY KEY, ts INTEGER, code TEXT, inspector TEXT, elecRehab TEXT, '
        'rehabDate TEXT, rehabUp TEXT, checkUp TEXT, notes TEXT'
    ),

    # Spares
    "spares": (
        'id TEXT PRIMARY KEY, ts INTEGER, itemType TEXT, "key" TEXT, model TEXT, '
        'partName TEXT, qty INTEGER, state TEXT, notes TEXT'
    ),
}

def _bootstrap() -> None:
    with db() as con:
        cur = con.cursor()
        # إنشاء الجداول
        for table, cols in SCHEMAS.items():
            cur.execute(f"CREATE TABLE IF NOT EXISTS {table} ({cols})")
        # فهارس مسرّعة
        cur.execute("CREATE INDEX IF NOT EXISTS idx_eng_supply_serial ON eng_supply(serial)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_eng_issue_serial ON eng_issue(serial)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_eng_rehab_serial ON eng_rehab(serial)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_eng_check_serial ON eng_check(serial)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_eng_upload_serial ON eng_upload(serial)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_eng_lathe_serial ON eng_lathe(serial)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_eng_pump_serial ON eng_pump(serial)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_eng_electrical_serial ON eng_electrical(serial)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_gen_supply_code ON gen_supply(code)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_gen_issue_code ON gen_issue(code)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_gen_inspect_code ON gen_inspect(code)")
        cur.execute('CREATE INDEX IF NOT EXISTS idx_spares_key ON spares("key")')
_bootstrap()

# ===================== Models =====================
class SyncItem(BaseModel):
    id: str
    store: str
    payload: Dict[str, Any]
    ts: int
    synced: Optional[bool] = None
    syncTs: Optional[int] = None

class SyncBatch(BaseModel):
    items: List[SyncItem]

class ExportBody(BaseModel):
    headers: List[str] = Field(..., description="عناوين الأعمدة بالعربية (ترتيب العرض)")
    rows: List[Any] = Field(..., description="List[Dict] أو List[List] (صف واحد لكل مفتاح)")
    filename: str = Field("alsami.xlsx", description="اسم الملف الناتج")
    sheet: str = Field("تقرير", description="اسم ورقة العمل")
    rtl: bool = Field(True, description="Right-To-Left")

# ===================== Helpers =====================
def _now_iso() -> str:
    return datetime.datetime.utcnow().isoformat() + "Z"

def _sanitize_filename(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]', "_", (name or "")).strip()
    return name or "alsami.xlsx"

def _assert_openpyxl() -> None:
    if Workbook is None:
        raise HTTPException(
            status_code=500,
            detail="openpyxl غير مثبت. ثبّت: pip install openpyxl==3.1.5",
        )

# ===================== Endpoints =====================
@app.get("/api/ping")
def ping() -> Dict[str, Any]:
    with db() as con:
        cur = con.cursor()
        counts = {}
        for t in SCHEMAS.keys():
            cur.execute(f"SELECT COUNT(*) n FROM {t}")
            counts[t] = cur.fetchone()["n"]
    return {
        "ok": True,
        "service": APP_TITLE,
        "version": APP_VER,
        "db": DB_PATH,
        "time": _now_iso(),
        "counts": counts,
    }

@app.get("/api/debug/counts")
def debug_counts() -> Dict[str, int]:
    with db() as con:
        cur = con.cursor()
        out: Dict[str, int] = {}
        for t in SCHEMAS:
            cur.execute(f"SELECT COUNT(*) n FROM {t}")
            out[t] = cur.fetchone()["n"]
        return out

@app.post("/api/sync/batch")
def sync_batch(body: SyncBatch) -> Dict[str, Any]:
    """
    UPsert مرن: يُدرج/يستبدل حسب أعمدة payload كما وردت.
    لو كانت payload فارغة نضمن id و ts على الأقل.
    """
    with db() as con:
        cur = con.cursor()
        for it in body.items:
            if it.store not in SCHEMAS:
                raise HTTPException(400, f"Unknown store '{it.store}'")

            # نحافظ على ترتيب المفاتيح كما وصل
            cols = list(json.loads(json.dumps(it.payload)).keys())
            if not cols:
                cols = ["id", "ts"]
                it.payload = {"id": it.id, "ts": it.ts}

            placeholders = ",".join("?" for _ in cols)
            sql = f"INSERT OR REPLACE INTO {it.store} ({','.join(cols)}) VALUES ({placeholders})"
            cur.execute(sql, [it.payload.get(c) for c in cols])
        con.commit()
    return {"ok": True, "count": len(body.items)}

@app.get("/api/search/{key}")
def api_search(key: str) -> Dict[str, Any]:
    """
    إرجاع كل العمليات المرتبطة بالمفتاح عبر كل الجداول (مرتبة بـ ts DESC).
    المفتاح للمحركات = serial، للمولدات = code، وللقطع = "key".
    """
    out: Dict[str, Any] = {}
    with db() as con:
        cur = con.cursor()

        def fetch(table: str, field: str) -> List[Dict[str, Any]]:
            cur.execute(f"SELECT * FROM {table} WHERE {field}=? ORDER BY ts DESC", (key,))
            return [dict(r) for r in cur.fetchall()]

        out["engines"] = {
            "supply": fetch("eng_supply", "serial"),
            "issue": fetch("eng_issue", "serial"),
            "rehab": fetch("eng_rehab", "serial"),
            "check": fetch("eng_check", "serial"),
            "upload": fetch("eng_upload", "serial"),
            "lathe": fetch("eng_lathe", "serial"),
            "pump": fetch("eng_pump", "serial"),
            "electrical": fetch("eng_electrical", "serial"),
        }
        out["generators"] = {
            "supply": fetch("gen_supply", "code"),
            "issue": fetch("gen_issue", "code"),
            "inspect": fetch("gen_inspect", "code"),
        }
        cur.execute('SELECT * FROM spares WHERE "key"=? ORDER BY ts DESC', (key,))
        out["spares"] = [dict(r) for r in cur.fetchall()]
    return out

@app.get("/api/last3/engines")
def last3_engines() -> Dict[str, Any]:
    with db() as con:
        cur = con.cursor()
        cur.execute("SELECT serial, prevSite, ts FROM eng_supply ORDER BY ts DESC LIMIT 3")
        rows = [dict(r) for r in cur.fetchall()]
    return {"items": rows}

@app.get("/api/last3/generators")
def last3_generators() -> Dict[str, Any]:
    with db() as con:
        cur = con.cursor()
        cur.execute("SELECT code, prevSite, ts FROM gen_supply ORDER BY ts DESC LIMIT 3")
        rows = [dict(r) for r in cur.fetchall()]
    return {"items": rows}

@app.post("/api/export/xlsx")
def export_xlsx(body: ExportBody = Body(...)) -> StreamingResponse:
    """
    تصدير Excel RTL ملوّن (Arabic RTL, zebra rows, borders, Tajawal font)
    - يقبل rows بالشكلين: List[Dict] أو List[List]
    - يلوّن العنوان + صفوف متبادلة، محاذاة يمين، تجميد رأس، اتجاه RTL
    """
    _assert_openpyxl()

    headers = list(body.headers or [])
    if not headers:
        raise HTTPException(400, "headers مطلوب")

    wb = Workbook()
    ws = wb.active
    ws.title = body.sheet or "تقرير"

    # RTL
    try:
        if body.rtl:
            ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    # أنماط
    header_fill = PatternFill("solid", fgColor="DCEAFE")  # أزرق فاتح
    zebra_fill  = PatternFill("solid", fgColor="F8FAFC")  # صفوف متبادلة
    head_font   = Font(name="Tajawal", bold=True, color="0F172A")
    cell_font   = Font(name="Tajawal")
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_r     = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # العناوين
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = head_font
        cell.border = border
        cell.alignment = align_r
        ws.column_dimensions[cell.column_letter].width = 22

    # البيانات
    r_index = 2
    for row in (body.rows or []):
        if isinstance(row, dict):
            values = [row.get(h, "") for h in headers]
        else:  # مصفوفة بالترتيب
            values = list(row) + [""] * (len(headers) - len(row))
            values = values[: len(headers)]
        ws.append(values)

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r_index, column=c)
            cell.font = cell_font
            cell.border = border
            cell.alignment = align_r
            if r_index % 2 == 0:
                cell.fill = zebra_fill
        r_index += 1

    # تجميد الصف الأول
    ws.freeze_panes = "A2"

       # إخراج
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = _sanitize_filename(body.filename or "alsami.xlsx")

    # 🔧 إصلاح الترميز (اسم الملف بالعربية)
    # نستخدم RFC 5987: filename*=UTF-8''<encoded_name>
    from urllib.parse import quote
    safe_filename = quote(filename)

    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}",
        "Access-Control-Expose-Headers": "Content-Disposition",
    }

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

# ===================== Serve Frontend (index.html) =====================
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse
import os

# تحديد المسار الصحيح لمجلد الواجهة الأمامية
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = os.path.join(BASE_DIR, "..", "frontend")

if os.path.isdir(FRONTEND_DIR):
    app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
else:
    print(f"⚠️ FRONTEND_DIR not found: {FRONTEND_DIR}")

@app.get("/")
def root_redirect():
    return RedirectResponse(url="/index.html")

# ============== نقطة تشغيل محلية اختيارية ==============
# شغّل: uvicorn main:app --reload --host 0.0.0.0 --port 9000
if __name__ == "__main__":  # pragma: no cover
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=9000)
